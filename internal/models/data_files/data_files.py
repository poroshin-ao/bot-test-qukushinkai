from internal.server.bot_init import bot
import openpyxl

async def download_file(file_path, destination):
    await bot.download_file(file_path, destination)

def parse_from_file(destination):
    data = dict()
    workbook = openpyxl.load_workbook(destination)
    worksheet = workbook.active

    ind = 1
    n = 0
    while n<2:
        name = worksheet.cell(row = ind, column = 1).value 
        
        if name == None:
            n = n+1
            ind = ind + 1
            continue
        n = 0

        data[str(name)] = list() 
         
        ind = ind + 1 
        while worksheet.cell(row = ind, column = 1).value != None:
            person = dict()
            person['FIO'] = str(worksheet.cell(row = ind, column = 1).value)
            person['num_fight'] = worksheet.cell(row = ind, column = 2).value
            data[name].append(person)
            ind = ind+1

    return data